"""Excel report and JSON audit log generation."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from .matcher import MatchResult


def _add_sheet(wb: Workbook, title: str, df: pd.DataFrame, header_fill: str = "DDEBF7") -> None:
    ws = wb.create_sheet(title=title)
    if df.empty:
        ws.append(["No records"])
        return
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        if r_idx == 1:
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor=header_fill)


def _summary_df(summary: dict[str, Any]) -> pd.DataFrame:
    return pd.DataFrame([{"Metric": k, "Value": v} for k, v in summary.items()])


def write_report(result: MatchResult, output_path: str | Path) -> None:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    _add_sheet(wb, "Matched", result.matched, "C6E0B4")
    _add_sheet(wb, "Unmatched Bank", result.unmatched_bank, "F8CBAD")
    _add_sheet(wb, "Unmatched Backend", result.unmatched_backend, "F8CBAD")
    _add_sheet(wb, "Partial Discrepant", result.partial, "FFE699")
    _add_sheet(wb, "Duplicates Bank", result.duplicates_bank, "B4C7E7")
    _add_sheet(wb, "Duplicates Backend", result.duplicates_backend, "B4C7E7")
    _add_sheet(wb, "Summary", _summary_df(result.summary), "D9D9D9")

    wb.save(output_path)

    audit_path = output_path.with_suffix(".audit_log.json")
    with open(audit_path, "w", encoding="utf-8") as f:
        json.dump(result.audit_log, f, indent=2, default=str)
